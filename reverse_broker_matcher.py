# -*- coding: utf-8 -*-
"""
尾盤均價線 (VWAP) 強勢股與主力分點逆向歸因引擎 (Reverse Broker Attribution Engine)
===================================================================================
核心功能：
1. 篩選當日「收盤高於全天均價線 (VWAP)、收盤位階接近全天最高 (收強)、且帶量」之強勢股。
2. 結合當日全市場券商分點 Parquet 資料，針對符合標的進行「分點買進均價逆向配對」：
   - 買進均價 >= VWAP 且緊貼收盤價 (證實主力在尾盤拉抬或高檔區間追價吃貨)。
   - 買進純度高 (排除當沖對敲，確認純吸籌)。
   - 買超張數與金額龐大。
3. 智慧判定主力畫像 (隔日沖主力 vs 波段外資/本土大戶)，並自動產出次日明確之交易指引。
4. 支援終端表格視覺化輸出與帶格式之 Excel (.xlsx) 決策報表產製。
"""

import os
import sys
import glob
import json
import time
import argparse
from typing import List, Dict, Optional, Tuple, Any

import duckdb
import pandas as pd
import numpy as np

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def load_broker_persona_db(config_path: Optional[str] = None) -> Dict[str, Any]:
    """載入主力券商畫像庫 (隔日沖/波段機構等)"""
    if not config_path:
        config_path = os.path.join(os.path.dirname(__file__), "broker_persona_db.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def classify_broker_persona(
    broker_id: str,
    broker_name: str,
    persona_db: Dict[str, Any]
) -> Tuple[str, str]:
    """
    根據券商代號與名稱判定主力屬性與次日建議策略
    回傳: (主力類型標籤, 次日作戰指引)
    """
    day_traders = persona_db.get("day_traders", {}).get("brokers", {})
    swing_inst = persona_db.get("swing_institutional", {}).get("brokers", {})

    b_id = str(broker_id).strip()
    
    # 隔日沖特徵檢查
    is_day_trader = (b_id in day_traders)
    if not is_day_trader:
        day_trade_keywords = ["台北", "建國", "土城永寧", "大安", "嘉義", "虎尾"]
        if any(kw in broker_name for kw in day_trade_keywords):
            is_day_trader = True

    if is_day_trader:
        tag = "⚡ 隔日沖急拉警報"
        action = "次日開高【絕對不追】；持股者 09:00~09:20 衝高滯漲宜【逢高停利】防主力倒貨。"
        return tag, action

    # 波段外資/機構特徵檢查
    is_swing = (b_id in swing_inst)
    if not is_swing:
        swing_keywords = ["摩根", "高盛", "瑞士信貸", "麥格理", "三多"]
        if any(kw in broker_name for kw in swing_keywords):
            is_swing = True

    if is_swing:
        tag = "💎 波段主力高位鎖碼"
        action = "趨勢具延續性；開盤平穩順勢抱牢，盤中拉回回測「主力買均價」不破可擇機加碼。"
        return tag, action

    # 預設本土實力派大戶
    tag = "🔥 本土實力大戶推升"
    action = "主力不計成本高位掃貨；次日以「當日 VWAP 均價線」為短線多空防守支撐。"
    return tag, action


def scan_tail_vwap_and_attribute(
    data_dir: str,
    target_date: Optional[str] = None,
    min_vol_sheets: float = 1000.0,
    min_turnover_yi: float = 0.3,
    min_vwap_premium: float = 0.5,
    max_vwap_premium: float = 6.0,
    min_broker_net_vol: float = 100.0,
    min_broker_net_amt_yi: float = 0.1,
    min_buy_purity: float = 70.0,
    max_close_diff_pct: float = 2.5,
    top_n: int = 50
) -> pd.DataFrame:
    """
    執行尾盤放量站上 VWAP 與分點逆向歸因主運算
    """
    # 支援單一整合目錄或分離快取目錄 (GitHub Actions 下載快取)
    search_dirs = [data_dir]
    for extra_d in ["./temp_cache_parquet", "./temp_cache_close", "./cloud_data"]:
        if os.path.exists(extra_d) and extra_d not in search_dirs:
            search_dirs.append(extra_d)

    raw_files = []
    for d in search_dirs:
        raw_files.extend(sorted(glob.glob(os.path.join(d, "*.parquet"))))

    if not raw_files:
        print(f"[!] 於目錄 {data_dir} 未找到任何 Parquet 檔案。")
        return pd.DataFrame()

    # 尋找目標日期的 close1 與 absr1
    close_files = [f for f in raw_files if "close1" in os.path.basename(f).lower()]
    absr_files = [f for f in raw_files if "absr1" in os.path.basename(f).lower() and "finmind" not in os.path.basename(f).lower()]

    if not close_files or not absr_files:
        print(f"[!] 缺少必要之 close1 (收盤日K) 或 absr1 (分點進出) 資料檔案。")
        return pd.DataFrame()

    if target_date:
        close_files = [f for f in close_files if target_date in os.path.basename(f)]
        absr_files = [f for f in absr_files if target_date in os.path.basename(f)]
        if not close_files or not absr_files:
            print(f"[!] 指定日期 {target_date} 查無完整的日K與分點資料。")
            return pd.DataFrame()
        target_close_file = close_files[-1]
        target_absr_file = absr_files[-1]
    else:
        target_close_file = close_files[-1]
        target_absr_file = absr_files[-1]

    # 解析實際執行的交易日期
    import re
    date_match = re.search(r"\d{4}-\d{2}-\d{2}", os.path.basename(target_close_file))
    actual_date = date_match.group(0) if date_match else "unknown"
    print(f"==================================================")
    print(f"[*] 尾盤放量站上 VWAP 與分點逆向歸因引擎啟動")
    print(f"[*] 分析交易日: {actual_date}")
    print(f"[*] 日K資料: {os.path.basename(target_close_file)}")
    print(f"[*] 分點資料: {os.path.basename(target_absr_file)}")
    print(f"[*] 標的門檻: 成交量 >= {min_vol_sheets:.0f}張, 金額 >= {min_turnover_yi:.1f}億, 均價溢價 {min_vwap_premium:.1f}%~{max_vwap_premium:.1f}%")
    print(f"[*] 主力門檻: 淨買超 >= {min_broker_net_vol:.0f}張, 淨買超 >= {min_broker_net_amt_yi:.2f}億, 純度 >= {min_buy_purity:.0f}%, 均價貼合度 <= {max_close_diff_pct:.1f}%")
    print(f"==================================================")

    # 載入名稱對照與畫像庫
    from find_similar_cases import get_stock_name_map, get_stock_market_map, get_broker_name_map
    stock_names = get_stock_name_map()
    stock_markets = get_stock_market_map()
    broker_names = get_broker_name_map()
    persona_db = load_broker_persona_db()

    close_sql_path = target_close_file.replace("\\", "/")
    absr_sql_path = target_absr_file.replace("\\", "/")

    sql = f"""
    WITH qualified_stocks AS (
        SELECT 
            symbol,
            name,
            market,
            close,
            open,
            high,
            low,
            volume / 1000.0 AS total_vol_sheets,
            turnover / 100000000.0 AS total_amt_yi,
            (turnover / volume) AS vwap,
            ((close - (turnover / volume)) / (turnover / volume)) * 100.0 AS vwap_premium_pct,
            CASE WHEN high > low THEN (close - low) / (high - low) ELSE 1.0 END AS pos_in_range,
            change,
            ((change / (close - change)) * 100.0) AS change_pct
        FROM read_parquet('{close_sql_path}')
        WHERE volume >= {min_vol_sheets * 1000.0}
          AND turnover >= {min_turnover_yi * 100000000.0}
          AND NOT (symbol LIKE '00%')
          AND symbol NOT IN ('ZZZZ', 'REG99', 'OTC99', 'Y9999')
          AND close > (turnover / volume)
          AND ((close - (turnover / volume)) / (turnover / volume)) * 100.0 BETWEEN {min_vwap_premium} AND {max_vwap_premium}
          AND (close >= open)
          AND CASE WHEN high > low THEN (close - low) / (high - low) ELSE 1.0 END >= 0.65
    ),
    broker_attribution AS (
        SELECT 
            s.symbol,
            s.name AS stock_name,
            s.market,
            s.close,
            s.vwap,
            s.vwap_premium_pct,
            s.change_pct,
            s.pos_in_range,
            s.total_vol_sheets,
            s.total_amt_yi,
            b.broker_id,
            b.buy_vol / 1000.0 AS broker_buy_vol_sheets,
            b.sell_vol / 1000.0 AS broker_sell_vol_sheets,
            b.net_vol / 1000.0 AS broker_net_vol_sheets,
            b.buy_amt / 100000.0 AS broker_buy_amt_yi,
            b.net_amt / 100000.0 AS broker_net_amt_yi,
            b.buy_avg_price AS broker_buy_avg_price,
            ((b.buy_avg_price - s.vwap) / s.vwap) * 100.0 AS buy_vs_vwap_pct,
            abs(b.buy_avg_price - s.close) / s.close * 100.0 AS buy_to_close_diff_pct,
            (b.buy_vol / (b.buy_vol + b.sell_vol)) * 100.0 AS buy_purity_pct,
            b.market_share
        FROM qualified_stocks s
        JOIN read_parquet('{absr_sql_path}') b ON s.symbol = b.symbol
        WHERE b.net_vol >= {min_broker_net_vol * 1000.0}
          AND (b.net_amt / 100000.0) >= {min_broker_net_amt_yi}
          AND b.buy_avg_price >= s.vwap * 0.995
          AND (b.buy_vol / (b.buy_vol + b.sell_vol)) * 100.0 >= {min_buy_purity}
    )
    SELECT * FROM broker_attribution
    WHERE buy_to_close_diff_pct <= {max_close_diff_pct}
    ORDER BY broker_net_amt_yi DESC
    """

    df = duckdb.query(sql).to_df()
    if df.empty:
        print("[!] 查無符合條件之標的與分點。")
        return df

    # 中文名稱映射與屬性判定
    df["股票名稱"] = df["symbol"].map(lambda x: stock_names.get(str(x), ""))
    # 若 close1 內已有名稱且非空白，則優先保留
    df["股票名稱"] = df.apply(lambda r: r["stock_name"] if pd.notna(r["stock_name"]) and r["stock_name"] != "" else r["股票名稱"], axis=1)
    df["市場別"] = df["symbol"].map(lambda x: stock_markets.get(str(x), "上市"))
    df["券商分點"] = df["broker_id"].map(lambda x: broker_names.get(str(x), str(x)))

    # 主力畫像與建議分類
    personas = []
    actions = []
    for _, row in df.iterrows():
        p_tag, p_act = classify_broker_persona(str(row["broker_id"]), str(row["券商分點"]), persona_db)
        personas.append(p_tag)
        actions.append(p_act)
    
    df["主力屬性"] = personas
    df["次日作戰指引"] = actions

    # 計算歸因綜合評分 (Attribution Score: 兼顧金額、純度與均價貼合度)
    # 分數 = 金額權重 + 純度獎勵 - 均價偏差懲罰
    df["歸因評分"] = (
        np.log1p(df["broker_net_amt_yi"].clip(lower=0)) * 25.0 +
        (df["buy_purity_pct"] - 70.0) * 1.0 -
        (df["buy_to_close_diff_pct"] * 8.0)
    ).round(1)

    # 欄位整理與格式化
    df["交易日期"] = actual_date
    df.sort_values(["歸因評分", "broker_net_amt_yi"], ascending=[False, False], inplace=True)
    if top_n and top_n > 0:
        df = df.head(top_n).reset_index(drop=True)

    return df


def export_to_excel(df: pd.DataFrame, output_path: str):
    """產出具備專業視覺格式之 Excel 決策報表"""
    if df.empty:
        return

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    
    # 挑選並重命名報表欄位
    display_cols = {
        "交易日期": "交易日期",
        "symbol": "股票代號",
        "股票名稱": "股票名稱",
        "市場別": "市場別",
        "close": "當日收盤價",
        "vwap": "全日均價(VWAP)",
        "vwap_premium_pct": "均價溢價率%",
        "change_pct": "當日漲幅%",
        "券商分點": "尾盤推手分點",
        "broker_buy_avg_price": "分點買均價",
        "buy_to_close_diff_pct": "均價貼合度%",
        "broker_net_vol_sheets": "分點淨買超(張)",
        "broker_net_amt_yi": "分點淨買超(億元)",
        "buy_purity_pct": "買進純度%",
        "歸因評分": "推升評分",
        "主力屬性": "主力屬性標籤",
        "次日作戰指引": "次日作戰指引"
    }
    
    out_df = df[list(display_cols.keys())].rename(columns=display_cols).copy()
    
    # 數值四捨五入美化
    out_df["當日收盤價"] = out_df["當日收盤價"].round(2)
    out_df["全日均價(VWAP)"] = out_df["全日均價(VWAP)"].round(2)
    out_df["均價溢價率%"] = out_df["均價溢價率%"].round(2)
    out_df["當日漲幅%"] = out_df["當日漲幅%"].round(2)
    out_df["分點買均價"] = out_df["分點買均價"].round(2)
    out_df["均價貼合度%"] = out_df["均價貼合度%"].round(2)
    out_df["分點淨買超(張)"] = out_df["分點淨買超(張)"].round(1)
    out_df["分點淨買超(億元)"] = out_df["分點淨買超(億元)"].round(3)
    out_df["買進純度%"] = out_df["買進純度%"].round(1)

    actual_output_path = output_path
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            out_df.to_excel(writer, index=False, sheet_name="尾盤放量站上VWAP決策表")
            
            # openpyxl 樣式微調
            workbook = writer.book
            worksheet = writer.sheets["尾盤放量站上VWAP決策表"]
            
            # 標題列美化
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9")
            )

            for col_idx, col_name in enumerate(out_df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                worksheet.column_dimensions[cell.column_letter].width = max(len(str(col_name)) * 2.6, 12)

            # 內容儲存格字型與對齊
            content_font = Font(name="微軟正黑體", size=10)
            for row in worksheet.iter_rows(min_row=2, max_row=len(out_df)+1, min_col=1, max_col=len(out_df.columns)):
                for cell in row:
                    cell.font = content_font
                    cell.border = thin_border
                    if cell.column in [1, 2, 4, 16]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif cell.column in [17]:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
    except PermissionError:
        actual_output_path = output_path.replace(".xlsx", f"_{int(time.time())}.xlsx")
        print(f"[!] 原檔案已被 Excel 開啟占用，改存至備用路徑: {actual_output_path}")
        with pd.ExcelWriter(actual_output_path, engine="openpyxl") as writer:
            out_df.to_excel(writer, index=False, sheet_name="尾盤放量站上VWAP決策表")

    print(f"[✓] 成功產出 Excel 決策報表: {actual_output_path}")
    return actual_output_path


def generate_tail_vwap_html_section(df: pd.DataFrame, top_n: int = 12) -> str:
    """生成現代 FinTech 響應式 HTML 郵件報告區塊 (相容 Gmail 行動端與電腦版)"""
    if df.empty:
        return ""

    actual_date = df["交易日期"].iloc[0] if "交易日期" in df.columns else ""
    sample_df = df.head(top_n)

    rows_html = ""
    for _, r in sample_df.iterrows():
        # 標籤顏色
        tag = str(r["主力屬性"])
        if "隔日沖" in tag:
            badge_style = "background-color: #fef2f2; color: #dc2626; border: 1px solid #fecaca;"
        elif "波段" in tag:
            badge_style = "background-color: #eff6ff; color: #2563eb; border: 1px solid #bfdbfe;"
        else:
            badge_style = "background-color: #fffbeb; color: #d97706; border: 1px solid #fde68a;"

        change_val = float(r["change_pct"])
        change_color = "#dc2626" if change_val > 0 else ("#16a34a" if change_val < 0 else "#4b5563")
        change_sign = "+" if change_val > 0 else ""

        rows_html += f"""
        <tr style="border-bottom: 1px solid #e2e8f0; font-size: 12px; height: 38px;">
            <td style="padding: 8px 6px; text-align: left; font-weight: 700; color: #0f172a;">
                {r['symbol']} {r['股票名稱']} <span style="font-size: 10px; color: #64748b; font-weight: normal;">({r['市場別']})</span>
            </td>
            <td style="padding: 8px 6px; text-align: right; color: {change_color}; font-weight: 700;">
                {r['close']:.1f} <span style="font-size: 10px;">({change_sign}{change_val:.1f}%)</span>
            </td>
            <td style="padding: 8px 6px; text-align: right; color: #334155;">
                {r['vwap']:.1f} <span style="font-size: 10px; color: #16a34a;">(+{r['vwap_premium_pct']:.1f}%)</span>
            </td>
            <td style="padding: 8px 6px; text-align: left; color: #1e293b;">
                <strong>{r['券商分點']}</strong>
                <div style="font-size: 10px; color: #64748b;">買均 {r['broker_buy_avg_price']:.1f} (貼合 {r['buy_to_close_diff_pct']:.1f}%)</div>
            </td>
            <td style="padding: 8px 6px; text-align: right; color: #0f172a;">
                <strong>+{r['broker_net_amt_yi']:.2f}億</strong>
                <div style="font-size: 10px; color: #64748b;">純度 {r['buy_purity_pct']:.0f}%</div>
            </td>
            <td style="padding: 8px 6px; text-align: center;">
                <span style="display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; {badge_style}">
                    {tag}
                </span>
            </td>
            <td style="padding: 8px 6px; text-align: left; font-size: 11px; color: #334155; line-height: 1.35;">
                {r['次日作戰指引']}
            </td>
        </tr>
        """

    html = f"""
    <div style="margin: 24px 0; border: 1px solid #cbd5e1; border-radius: 10px; overflow: hidden; background: #ffffff; box-shadow: 0 2px 6px rgba(0,0,0,0.04);">
        <div style="background: linear-gradient(135deg, #1e3a8a 0%, #1e40af 100%); color: #ffffff; padding: 14px 18px;">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <div>
                    <h3 style="margin: 0; font-size: 16px; font-weight: 800; letter-spacing: 0.5px;">🎯 尾盤放量站上 VWAP 與主力分點逆向歸因雷達</h3>
                    <p style="margin: 4px 0 0; font-size: 11px; color: #bfdbfe; line-height: 1.4;">
                        精準捕捉當日收盤強勢站穩 VWAP 均價線且高位掃貨之標的，結合分點買進均價逆向定位隔日沖與波段主力。
                    </p>
                </div>
                <div style="background: rgba(255,255,255,0.15); border-radius: 6px; padding: 4px 10px; font-size: 11px; font-weight: 700;">
                    基準日: {actual_date}
                </div>
            </div>
        </div>
        
        <div style="overflow-x: auto; padding: 6px 12px 14px;">
            <table style="width: 100%; border-collapse: collapse; text-align: left; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;">
                <thead>
                    <tr style="border-bottom: 2px solid #cbd5e1; background-color: #f8fafc; font-size: 11px; color: #475569; height: 34px;">
                        <th style="padding: 6px; text-align: left;">股票標的</th>
                        <th style="padding: 6px; text-align: right;">收盤 (漲跌)</th>
                        <th style="padding: 6px; text-align: right;">VWAP均價 (溢價)</th>
                        <th style="padding: 6px; text-align: left;">尾盤推手分點</th>
                        <th style="padding: 6px; text-align: right;">淨買超 (純度)</th>
                        <th style="padding: 6px; text-align: center;">主力屬性</th>
                        <th style="padding: 6px; text-align: left;">次日開盤作戰指引</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
        </div>
        <div style="background-color: #f8fafc; padding: 8px 16px; border-top: 1px solid #e2e8f0; font-size: 11px; color: #64748b; text-align: right;">
            💡 完整 30+ 檔標的之詳細均價貼合度與推升評分請參閱隨信 Excel 附件。
        </div>
    </div>
    """
    return html


def append_tail_vwap_sheet_to_excel(excel_path: str, tail_df: pd.DataFrame):
    """將尾盤放量站上 VWAP 決策表追加至指定的 Excel 檔案中"""
    if tail_df.empty or not os.path.exists(excel_path):
        return

    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = "尾盤放量站上VWAP決策表"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(title=sheet_name)
    
    display_cols = {
        "交易日期": "交易日期",
        "symbol": "股票代號",
        "股票名稱": "股票名稱",
        "市場別": "市場別",
        "close": "當日收盤價",
        "vwap": "全日均價(VWAP)",
        "vwap_premium_pct": "均價溢價率%",
        "change_pct": "當日漲幅%",
        "券商分點": "尾盤推手分點",
        "broker_buy_avg_price": "分點買均價",
        "buy_to_close_diff_pct": "均價貼合度%",
        "broker_net_vol_sheets": "分點淨買超(張)",
        "broker_net_amt_yi": "分點淨買超(億元)",
        "buy_purity_pct": "買進純度%",
        "歸因評分": "推升評分",
        "主力屬性": "主力屬性標籤",
        "次日作戰指引": "次日作戰指引"
    }
    
    out_df = tail_df[list(display_cols.keys())].rename(columns=display_cols).copy()
    out_df["當日收盤價"] = out_df["當日收盤價"].round(2)
    out_df["全日均價(VWAP)"] = out_df["全日均價(VWAP)"].round(2)
    out_df["均價溢價率%"] = out_df["均價溢價率%"].round(2)
    out_df["當日漲幅%"] = out_df["當日漲幅%"].round(2)
    out_df["分點買均價"] = out_df["分點買均價"].round(2)
    out_df["均價貼合度%"] = out_df["均價貼合度%"].round(2)
    out_df["分點淨買超(張)"] = out_df["分點淨買超(張)"].round(1)
    out_df["分點淨買超(億元)"] = out_df["分點淨買超(億元)"].round(3)
    out_df["買進純度%"] = out_df["買進純度%"].round(1)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # 寫入標題列
    headers = list(out_df.columns)
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = max(len(str(col_name)) * 2.6, 12)

    # 寫入資料列
    content_font = Font(name="微軟正黑體", size=10)
    for row_idx, row_data in enumerate(out_df.values, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = content_font
            cell.border = thin_border
            if col_idx in [1, 2, 4, 16]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [17]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    wb.save(excel_path)
    print(f"[✓] 成功將尾盤放量工作表追加至 Excel: {excel_path}")


def send_tail_vwap_report_email(
    df: pd.DataFrame,
    output_excel: Optional[str] = None,
    recipients: Optional[List[str]] = None
) -> bool:
    """產出完整獨立 Email 並發送"""
    if df.empty:
        print("[!] 查無資料，跳過 Email 發送。")
        return False

    from send_email_report import send_email_report, send_telegram_notify

    actual_date = df["交易日期"].iloc[0]
    section_html = generate_tail_vwap_html_section(df, top_n=15)

    full_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>🎯 台股尾盤放量站上 VWAP 與分點逆向歸因日報 ({actual_date})</title>
    </head>
    <body style="margin: 0; padding: 20px; background-color: #f1f5f9; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;">
        <div style="max-width: 960px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.06);">
            <!-- Header -->
            <div style="background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%); padding: 24px 30px; color: #ffffff;">
                <h1 style="margin: 0; font-size: 22px; font-weight: 900; letter-spacing: 0.5px;">
                    🎯 台股尾盤放量站上 VWAP 與分點逆向歸因日報
                </h1>
                <p style="margin: 6px 0 0; font-size: 13px; color: #94a3b8;">
                    交易基準日: {actual_date} ｜ 涵蓋隔日沖獵殺警報、外資波段鎖碼與本土主力動態
                </p>
            </div>
            
            <!-- Body -->
            <div style="padding: 20px 24px;">
                {section_html}
                
                <div style="margin-top: 20px; padding: 14px; background-color: #f8fafc; border-left: 4px solid #3b82f6; border-radius: 4px; font-size: 12px; color: #334155; line-height: 1.6;">
                    <strong>💡 次日開盤實戰重點：</strong><br>
                    1. <strong>⚡ 隔日沖急拉警報</strong>：若標的早盤開高，切忌追價；持股者可在 09:00~09:20 衝高動能減緩時分批獲利了結。<br>
                    2. <strong>💎 波段主力鎖碼</strong>：主力承擔隔夜成本推升，盤中若回測「主力買均價」或「當日 VWAP」有撐，為極佳之右側佈局點。<br>
                    3. 本郵件已隨信附上完整 Excel 決策報表，歡迎開啟進行深度個股檢視。
                </div>
            </div>
            
            <!-- Footer -->
            <div style="background-color: #f8fafc; padding: 14px 24px; text-align: center; font-size: 11px; color: #94a3b8; border-top: 1px solid #e2e8f0;">
                台股券商分點量化分析系統自動產製 ｜ 本日報僅供量化策略研究，不構成任何投資買賣建議。
            </div>
        </div>
    </body>
    </html>
    """

    subject = f"🎯 台股尾盤放量站上 VWAP 與分點逆向歸因日報 ({actual_date}) | 隔日沖獵殺 ＋ 波段鎖碼"
    attachments = [output_excel] if output_excel and os.path.exists(output_excel) else None

    success = send_email_report(
        subject=subject,
        html_content=full_html,
        recipients=recipients,
        attachment_paths=attachments
    )

    if success:
        tg_msg = (
            f"🎯 *台股尾盤放量站上 VWAP 歸因日報 ({actual_date})*\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📊 篩選強勢標的：`{len(df)} 檔`\n"
            f"⚡ 隔日沖警報：`{len(df[df['主力屬性'].str.contains('隔日沖')])} 檔`\n"
            f"💎 波段主力鎖碼：`{len(df[df['主力屬性'].str.contains('波段')])} 檔`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📧 專屬 HTML 郵件與 Excel 報表已成功寄出！"
        )
        send_telegram_notify(tg_msg)

    return success


def print_summary_table(df: pd.DataFrame):
    """於終端印出格式化繁體中文摘要"""
    if df.empty:
        return

    print("\n" + "=" * 115)
    print(f"{'代號':<6} {'名稱':<8} {'收盤':>7} {'VWAP':>7} {'推手分點':<12} {'買均價':>7} {'買超(張)':>8} {'金額(億)':>7} {'純度%':>6} {'主力屬性':<14} {'推升評分':>7}")
    print("-" * 115)

    for _, r in df.iterrows():
        print(f"{str(r['symbol']):<6} {str(r['股票名稱'])[:6]:<8} {r['close']:>7.1f} {r['vwap']:>7.1f} {str(r['券商分點'])[:8]:<12} {r['broker_buy_avg_price']:>7.1f} {r['broker_net_vol_sheets']:>8.1f} {r['broker_net_amt_yi']:>7.2f} {r['buy_purity_pct']:>6.1f} {str(r['主力屬性']):<14} {r['歸因評分']:>7.1f}")
    print("=" * 115 + "\n")


def main():
    parser = argparse.ArgumentParser(description="尾盤均價線 (VWAP) 強勢股與分點逆向歸因引擎")
    parser.add_argument("--data-dir", default=r"d:\MyProject\stock_data_analysis\20260822分點資料", help="Parquet 資料目錄")
    parser.add_argument("--date", default=None, help="指定分析日期 (YYYY-MM-DD，預設取最新日期)")
    parser.add_argument("--min-vol-sheets", type=float, default=1000.0, help="最小成交張數 (預設 1000 張)")
    parser.add_argument("--min-turnover-yi", type=float, default=0.3, help="最小成交金額 (億元，預設 0.3 億)")
    parser.add_argument("--min-vwap-premium", type=float, default=0.5, help="最小收盤高於均價百分比 (預設 0.5%)")
    parser.add_argument("--max-vwap-premium", type=float, default=6.0, help="最大收盤高於均價百分比 (預設 6.0%)")
    parser.add_argument("--min-broker-net-vol", type=float, default=100.0, help="主力最小淨買超張數 (預設 100 張)")
    parser.add_argument("--min-broker-net-amt-yi", type=float, default=0.1, help="主力最小淨買超金額 (億元，預設 0.1 億)")
    parser.add_argument("--min-buy-purity", type=float, default=70.0, help="主力買進純度% (預設 70%)")
    parser.add_argument("--top-n", type=int, default=30, help="輸出前 N 檔結果")
    parser.add_argument("--output-excel", default=None, help="匯出 Excel 報表路徑 (選填)")
    parser.add_argument("--send-email", action="store_true", help="產出後自動發送 Email 報告")

    args = parser.parse_args()

    df = scan_tail_vwap_and_attribute(
        data_dir=args.data_dir,
        target_date=args.date,
        min_vol_sheets=args.min_vol_sheets,
        min_turnover_yi=args.min_turnover_yi,
        min_vwap_premium=args.min_vwap_premium,
        max_vwap_premium=args.max_vwap_premium,
        min_broker_net_vol=args.min_broker_net_vol,
        min_broker_net_amt_yi=args.min_broker_net_amt_yi,
        min_buy_purity=args.min_buy_purity,
        top_n=args.top_n
    )

    if not df.empty:
        print_summary_table(df)
        
        actual_date = df["交易日期"].iloc[0]
        output_excel = args.output_excel
        if not output_excel:
            output_excel = os.path.join(
                os.path.dirname(__file__), 
                "output", 
                f"尾盤放量站上VWAP主力歸因表_{actual_date}.xlsx"
            )
        output_excel = export_to_excel(df, output_excel)

        if args.send_email:
            send_tail_vwap_report_email(df, output_excel=output_excel)


if __name__ == "__main__":
    main()
