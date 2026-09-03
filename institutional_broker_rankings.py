# -*- coding: utf-8 -*-
"""
外資與本土法人分點買賣超解密引擎 (Institutional Broker Rankings Engine)
======================================================================
核心功能：
1. 外資席位買賣超解密：
   - 聚合外商券商席位 (台灣摩根士丹利、摩根大通、高盛、美林、麥格理、瑞銀等)
   - 精確拆解各外資席位當日重押個股、淨買超金額、買進均價與純度
   - 一眼看穿「假外資隔日沖 (如美林)」與「真波段長莊 (如大摩、高盛)」的資金流向
2. 本土法人/投信大戶重押解密：
   - 聚合本土券商法人部 (如國票-敦北法人) 與各大券商總公司專戶
   - 捕捉國內基金經理人與主力代操集團當日同步重押之核心飆股
3. 全市場分點多空之王排行：
   - 全市場單日做多銀彈最猛烈之 TOP 10 分點
   - 全市場單日出貨調節最大之 TOP 10 分點
4. 匯出多工作表專業 Excel (.xlsx) 決策報表，並支援自動渲染 HTML 郵件派發。
"""

import os
import sys
import glob
import json
import time
import argparse
from typing import List, Dict, Optional, Tuple, Any

import duckdb
import pandas as pd
import numpy as np

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# 導入名稱對照
from find_similar_cases import get_stock_name_map, get_stock_market_map, get_broker_name_map

# 外商券商在台席位清單
FOREIGN_BROKERS = {
    "1470": "台灣摩根士丹利",
    "8440": "摩根大通",
    "1480": "美商高盛",
    "1440": "美林",
    "1360": "港商麥格理",
    "1380": "新加坡商瑞銀",
    "1650": "新加坡商瑞銀",
    "1560": "港商野村",
    "1520": "瑞士信貸",
    "1590": "美商花旗",
    "1350": "港商法興",
    "1160": "日商大和"
}

# 本土法人業務部與總公司專戶清單 (國內投信/代操/壽險重鎮)
DOMESTIC_INST_BROKERS = {
    "779c": "國票-敦北法人",
    "9200": "凱基總公司",
    "9600": "富邦總公司",
    "9800": "元大總公司",
    "9A00": "永豐金總公司",
    "9100": "群益金鼎總公司",
    "8880": "國泰總公司",
    "5850": "統一總公司",
    "7000": "兆豐總公司",
    "1020": "合庫總公司",
    "5920": "元富總公司",
    "6160": "中信總公司",
    "9B00": "台新總公司"
}


def run_institutional_ranking_analysis(
    data_dir: str,
    target_date: Optional[str] = None,
    min_foreign_net_amt_yi: float = 0.5,  # 外資單股淨買超門檻 (億元，預設 0.5 億)
    min_inst_net_amt_yi: float = 0.2,     # 本土法人單股淨買超門檻 (億元，預設 0.2 億)
    top_n: int = 15
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    """
    執行三大維度分點買賣超排行運算：
    回傳 (外資重押榜, 本土法人重押榜, 全市場多空之王榜, 實際交易日期)
    """
    search_dirs = [data_dir]
    for d in ["./temp_cache_parquet", "./cloud_data", "./20260822分點資料"]:
        if os.path.exists(d) and d not in search_dirs:
            search_dirs.append(d)

    raw_files = []
    for d in search_dirs:
        raw_files.extend(sorted(glob.glob(os.path.join(d, "*.parquet"))))

    absr_files = [
        f for f in raw_files 
        if "absr1" in os.path.basename(f).lower() and "finmind" not in os.path.basename(f).lower()
    ]

    if not absr_files:
        print(f"[!] 於目錄未找到任何 absr1 分點 Parquet 檔案。")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), ""

    if target_date:
        absr_files = [f for f in absr_files if target_date in os.path.basename(f)]
        if not absr_files:
            print(f"[!] 指定日期 {target_date} 查無分點資料。")
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), ""
        target_file = absr_files[-1]
    else:
        target_file = absr_files[-1]

    import re
    date_match = re.search(r"\d{4}-\d{2}-\d{2}", os.path.basename(target_file))
    actual_date = date_match.group(0) if date_match else "unknown"

    print("=" * 60)
    print(f"[*] 外資與本土法人分點買賣超解密引擎啟動")
    print(f"[*] 分析交易日: {actual_date}")
    print(f"[*] 分點資料來源: {os.path.basename(target_file)}")
    print("=" * 60)

    stock_names = get_stock_name_map()
    stock_markets = get_stock_market_map()
    broker_names = get_broker_name_map()
    sql_path = target_file.replace("\\", "/")

    # 1. 外資席位當日重押榜
    foreign_ids_sql = ",".join([f"'{k}'" for k in FOREIGN_BROKERS.keys()])
    sql_foreign = f"""
        SELECT 
            b.broker_id,
            b.symbol,
            b.buy_vol / 1000.0 AS buy_sheets,
            b.sell_vol / 1000.0 AS sell_sheets,
            b.net_vol / 1000.0 AS net_sheets,
            b.buy_amt / 100000.0 AS buy_amt_yi,
            b.sell_amt / 100000.0 AS sell_amt_yi,
            b.net_amt / 100000.0 AS net_amt_yi,
            b.buy_avg_price,
            (b.buy_vol / (b.buy_vol + b.sell_vol)) * 100.0 AS buy_purity_pct,
            b.market_share
        FROM read_parquet('{sql_path}') b
        WHERE b.broker_id IN ({foreign_ids_sql})
          AND (b.net_amt / 100000.0) >= {min_foreign_net_amt_yi}
          AND NOT (b.symbol LIKE '00%')
          AND b.symbol NOT IN ('ZZZZ', 'REG99', 'OTC99', 'Y9999')
        ORDER BY b.net_amt DESC
        LIMIT {top_n}
    """
    df_foreign = duckdb.query(sql_foreign).df()
    if not df_foreign.empty:
        df_foreign["券商分點"] = df_foreign["broker_id"].map(lambda x: FOREIGN_BROKERS.get(str(x), broker_names.get(str(x), str(x))))
        df_foreign["股票代號"] = df_foreign["symbol"]
        df_foreign["股票名稱"] = df_foreign["symbol"].map(lambda x: stock_names.get(str(x), str(x)))
        df_foreign["市場別"] = df_foreign["symbol"].map(lambda x: stock_markets.get(str(x), "上市"))
        # 標註屬性
        df_foreign["外資屬性"] = df_foreign["broker_id"].map(
            lambda x: "⚡ 短線高頻/隔日沖" if str(x) in ["1440"] else "💎 波段機構主力"
        )

    # 2. 本土法人/總公司專戶重押榜
    inst_ids_sql = ",".join([f"'{k}'" for k in DOMESTIC_INST_BROKERS.keys()])
    sql_inst = f"""
        SELECT 
            b.broker_id,
            b.symbol,
            b.buy_vol / 1000.0 AS buy_sheets,
            b.sell_vol / 1000.0 AS sell_sheets,
            b.net_vol / 1000.0 AS net_sheets,
            b.buy_amt / 100000.0 AS buy_amt_yi,
            b.sell_amt / 100000.0 AS sell_amt_yi,
            b.net_amt / 100000.0 AS net_amt_yi,
            b.buy_avg_price,
            (b.buy_vol / (b.buy_vol + b.sell_vol)) * 100.0 AS buy_purity_pct,
            b.market_share
        FROM read_parquet('{sql_path}') b
        WHERE b.broker_id IN ({inst_ids_sql})
          AND (b.net_amt / 100000.0) >= {min_inst_net_amt_yi}
          AND NOT (b.symbol LIKE '00%')
          AND b.symbol NOT IN ('ZZZZ', 'REG99', 'OTC99', 'Y9999')
        ORDER BY b.net_amt DESC
        LIMIT {top_n}
    """
    df_inst = duckdb.query(sql_inst).df()
    if not df_inst.empty:
        df_inst["券商分點"] = df_inst["broker_id"].map(lambda x: DOMESTIC_INST_BROKERS.get(str(x), broker_names.get(str(x), str(x))))
        df_inst["股票代號"] = df_inst["symbol"]
        df_inst["股票名稱"] = df_inst["symbol"].map(lambda x: stock_names.get(str(x), str(x)))
        df_inst["市場別"] = df_inst["symbol"].map(lambda x: stock_markets.get(str(x), "上市"))
        df_inst["法人標籤"] = df_inst["buy_purity_pct"].map(
            lambda x: "🎯 絕對鎖碼 (純買無賣)" if x >= 95.0 else ("🔥 積極建倉" if x >= 75.0 else "⚖️ 換手進出")
        )

    # 3. 全市場分點多空之王排行 (全市場今日淨買超金額 Top 10 與淨賣超 Top 10)
    sql_bull = f"""
        SELECT 
            broker_id,
            SUM(buy_amt) / 100000.0 AS total_buy_yi,
            SUM(sell_amt) / 100000.0 AS total_sell_yi,
            SUM(net_amt) / 100000.0 AS net_amt_yi,
            SUM(buy_vol) / 1000.0 AS total_buy_sheets,
            SUM(sell_vol) / 1000.0 AS total_sell_sheets,
            SUM(net_vol) / 1000.0 AS net_vol_sheets
        FROM read_parquet('{sql_path}')
        GROUP BY broker_id
        ORDER BY net_amt_yi DESC
        LIMIT 10
    """
    df_bull = duckdb.query(sql_bull).df()
    df_bull["多空陣營"] = "🐂 多頭司令 (買超之王)"
    df_bull["券商分點"] = df_bull["broker_id"].map(lambda x: broker_names.get(str(x), str(x)))

    sql_bear = f"""
        SELECT 
            broker_id,
            SUM(buy_amt) / 100000.0 AS total_buy_yi,
            SUM(sell_amt) / 100000.0 AS total_sell_yi,
            SUM(net_amt) / 100000.0 AS net_amt_yi,
            SUM(buy_vol) / 1000.0 AS total_buy_sheets,
            SUM(sell_vol) / 1000.0 AS total_sell_sheets,
            SUM(net_vol) / 1000.0 AS net_vol_sheets
        FROM read_parquet('{sql_path}')
        GROUP BY broker_id
        ORDER BY net_amt_yi ASC
        LIMIT 10
    """
    df_bear = duckdb.query(sql_bear).df()
    df_bear["多空陣營"] = "🐻 空頭殺手 (賣超調節)"
    df_bear["券商分點"] = df_bear["broker_id"].map(lambda x: broker_names.get(str(x), str(x)))

    df_market_kings = pd.concat([df_bull, df_bear], ignore_index=True)

    return df_foreign, df_inst, df_market_kings, actual_date


def export_institutional_rankings_to_excel(
    df_foreign: pd.DataFrame,
    df_inst: pd.DataFrame,
    df_market_kings: pd.DataFrame,
    output_path: str
) -> str:
    """匯出專業 Excel 報表 (含三個工作表)"""
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    actual_output_path = output_path

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    content_font = Font(name="微軟正黑體", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    def _style_sheet(ws, fill_color: str, df_cols: List[str]):
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx, col_name in enumerate(df_cols, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            ws.column_dimensions[cell.column_letter].width = max(len(str(col_name)) * 2.8, 12)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(df_cols)):
            for cell in row:
                cell.font = content_font
                cell.border = thin_border
                cell.alignment = center_align if cell.column in [1, 2, 4] else right_align

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Sheet 1: 外資席位重押榜
            if not df_foreign.empty:
                cols_f = ["券商分點", "股票代號", "股票名稱", "市場別", "外資屬性", "net_amt_yi", "buy_avg_price", "net_sheets", "buy_purity_pct"]
                labels_f = ["外資分點", "股票代號", "股票名稱", "市場別", "外資屬性", "淨買超(億元)", "買進均價", "淨買超(張)", "買進純度%"]
                out_f = df_foreign[cols_f].rename(columns=dict(zip(cols_f, labels_f)))
                out_f["淨買超(億元)"] = out_f["淨買超(億元)"].round(2)
                out_f["買進均價"] = out_f["買進均價"].round(2)
                out_f["淨買超(張)"] = out_f["淨買超(張)"].round(1)
                out_f["買進純度%"] = out_f["買進純度%"].round(1)
                out_f.to_excel(writer, index=False, sheet_name="外資主要席位重押榜")
                _style_sheet(writer.sheets["外資主要席位重押榜"], "1E3A8A", labels_f)

            # Sheet 2: 本土法人部重押榜
            if not df_inst.empty:
                cols_i = ["券商分點", "股票代號", "股票名稱", "市場別", "法人標籤", "net_amt_yi", "buy_avg_price", "net_sheets", "buy_purity_pct"]
                labels_i = ["本土法人分點", "股票代號", "股票名稱", "市場別", "鎖碼特徵", "淨買超(億元)", "買進均價", "淨買超(張)", "買進純度%"]
                out_i = df_inst[cols_i].rename(columns=dict(zip(cols_i, labels_i)))
                out_i["淨買超(億元)"] = out_i["淨買超(億元)"].round(2)
                out_i["買進均價"] = out_i["買進均價"].round(2)
                out_i["淨買超(張)"] = out_i["淨買超(張)"].round(1)
                out_i["買進純度%"] = out_i["買進純度%"].round(1)
                out_i.to_excel(writer, index=False, sheet_name="本土法人與總公司重押榜")
                _style_sheet(writer.sheets["本土法人與總公司重押榜"], "B45309", labels_i)

            # Sheet 3: 全市場分點多空榜
            if not df_market_kings.empty:
                cols_m = ["多空陣營", "券商分點", "net_amt_yi", "total_buy_yi", "total_sell_yi", "net_vol_sheets"]
                labels_m = ["多空陣營", "券商分點", "淨買超金額(億元)", "總買進金額(億元)", "總賣出金額(億元)", "淨買超張數(張)"]
                out_m = df_market_kings[cols_m].rename(columns=dict(zip(cols_m, labels_m)))
                out_m["淨買超金額(億元)"] = out_m["淨買超金額(億元)"].round(2)
                out_m["總買進金額(億元)"] = out_m["總買進金額(億元)"].round(2)
                out_m["總賣出金額(億元)"] = out_m["總賣出金額(億元)"].round(2)
                out_m["淨買超張數(張)"] = out_m["淨買超張數(張)"].round(1)
                out_m.to_excel(writer, index=False, sheet_name="全市場分點多空之王")
                _style_sheet(writer.sheets["全市場分點多空之王"], "047857", labels_m)

    except PermissionError:
        actual_output_path = output_path.replace(".xlsx", f"_{int(time.time())}.xlsx")
        print(f"[!] 原檔案已被開啟占用，改存至備用路徑: {actual_output_path}")
        with pd.ExcelWriter(actual_output_path, engine="openpyxl") as writer:
            if not df_foreign.empty:
                df_foreign.to_excel(writer, index=False, sheet_name="外資主要席位重押榜")
            if not df_inst.empty:
                df_inst.to_excel(writer, index=False, sheet_name="本土法人與總公司重押榜")
            if not df_market_kings.empty:
                df_market_kings.to_excel(writer, index=False, sheet_name="全市場分點多空之王")

    print(f"[✓] 成功產出 Excel 報表: {actual_output_path}")
    return actual_output_path


def send_institutional_rankings_email(
    df_foreign: pd.DataFrame,
    df_inst: pd.DataFrame,
    df_market_kings: pd.DataFrame,
    actual_date: str,
    output_excel: Optional[str] = None,
    recipients: Optional[List[str]] = None
) -> bool:
    """產出專屬 HTML 郵件並發送"""
    from send_email_report import send_email_report, send_telegram_notify

    # 外資表 Rows
    foreign_rows = ""
    for _, r in df_foreign.head(10).iterrows():
        is_day_trade = ("短線" in str(r["外資屬性"]))
        badge_style = "background-color: #fef2f2; color: #dc2626; border: 1px solid #fecaca;" if is_day_trade else "background-color: #eff6ff; color: #2563eb; border: 1px solid #bfdbfe;"
        foreign_rows += f"""
        <tr style="border-bottom: 1px solid #e2e8f0; font-size: 12px; height: 36px;">
            <td style="padding: 6px 8px; font-weight: 700; color: #1e3a8a;">{r['券商分點']}</td>
            <td style="padding: 6px 8px; font-weight: 700; color: #0f172a;">
                {r['股票代號']} {r['股票名稱']} <span style="font-size: 10px; color: #64748b;">({r['市場別']})</span>
            </td>
            <td style="padding: 6px 8px; text-align: right; color: #dc2626; font-weight: 700;">+{r['net_amt_yi']:.2f} 億</td>
            <td style="padding: 6px 8px; text-align: right; color: #334155;">{r['buy_avg_price']:.1f} 元</td>
            <td style="padding: 6px 8px; text-align: right; color: #334155;">純度 {r['buy_purity_pct']:.0f}%</td>
            <td style="padding: 6px 8px; text-align: center;">
                <span style="display: inline-block; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: 700; {badge_style}">
                    {r['外資屬性']}
                </span>
            </td>
        </tr>
        """

    # 本土法人表 Rows
    inst_rows = ""
    for _, r in df_inst.head(10).iterrows():
        inst_rows += f"""
        <tr style="border-bottom: 1px solid #e2e8f0; font-size: 12px; height: 36px;">
            <td style="padding: 6px 8px; font-weight: 700; color: #b45309;">{r['券商分點']}</td>
            <td style="padding: 6px 8px; font-weight: 700; color: #0f172a;">
                {r['股票代號']} {r['股票名稱']} <span style="font-size: 10px; color: #64748b;">({r['市場別']})</span>
            </td>
            <td style="padding: 6px 8px; text-align: right; color: #dc2626; font-weight: 700;">+{r['net_amt_yi']:.2f} 億</td>
            <td style="padding: 6px 8px; text-align: right; color: #334155;">{r['buy_avg_price']:.1f} 元</td>
            <td style="padding: 6px 8px; text-align: right; color: #334155;">純度 {r['buy_purity_pct']:.0f}%</td>
            <td style="padding: 6px 8px; text-align: center;">
                <span style="display: inline-block; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: 700; background-color: #fef3c7; color: #92400e; border: 1px solid #fde68a;">
                    {r['法人標籤']}
                </span>
            </td>
        </tr>
        """

    # 全市場多空之王 Rows
    bull_df = df_market_kings[df_market_kings["多空陣營"].str.contains("多頭")].head(5)
    bear_df = df_market_kings[df_market_kings["多空陣營"].str.contains("空頭")].head(5)
    kings_rows = ""
    for _, r in bull_df.iterrows():
        kings_rows += f"""
        <tr style="border-bottom: 1px solid #e2e8f0; font-size: 12px; height: 32px;">
            <td style="padding: 6px 8px; color: #16a34a; font-weight: 700;">🐂 {r['券商分點']}</td>
            <td style="padding: 6px 8px; text-align: right; font-weight: 700; color: #dc2626;">+{r['net_amt_yi']:.2f} 億</td>
            <td style="padding: 6px 8px; text-align: right; color: #64748b;">買進 {r['total_buy_yi']:.1f}億 / 賣出 {r['total_sell_yi']:.1f}億</td>
        </tr>
        """
    for _, r in bear_df.iterrows():
        kings_rows += f"""
        <tr style="border-bottom: 1px solid #e2e8f0; font-size: 12px; height: 32px;">
            <td style="padding: 6px 8px; color: #dc2626; font-weight: 700;">🐻 {r['券商分點']}</td>
            <td style="padding: 6px 8px; text-align: right; font-weight: 700; color: #16a34a;">{r['net_amt_yi']:.2f} 億</td>
            <td style="padding: 6px 8px; text-align: right; color: #64748b;">買進 {r['total_buy_yi']:.1f}億 / 賣出 {r['total_sell_yi']:.1f}億</td>
        </tr>
        """

    full_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>🏛️ 台股外資與本土法人分點買賣超解密日報 ({actual_date})</title>
    </head>
    <body style="margin: 0; padding: 20px; background-color: #f1f5f9; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;">
        <div style="max-width: 980px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.06);">
            
            <!-- Header -->
            <div style="background: linear-gradient(135deg, #1e3a8a 0%, #0f172a 100%); padding: 26px 30px; color: #ffffff;">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div>
                        <h1 style="margin: 0; font-size: 22px; font-weight: 900; letter-spacing: 0.5px;">
                            🏛️ 台股外資與本土法人分點買賣超解密日報
                        </h1>
                        <p style="margin: 6px 0 0; font-size: 13px; color: #bfdbfe;">
                            交易基準日: {actual_date} ｜ 穿透三大法人總計表象，直擊各大券商席位與操盤本質
                        </p>
                    </div>
                </div>
            </div>
            
            <div style="padding: 24px;">
                
                <!-- 核心洞察卡片 -->
                <div style="margin-bottom: 22px; padding: 14px 18px; background-color: #eff6ff; border-left: 4px solid #3b82f6; border-radius: 6px; font-size: 12px; color: #1e3a8a; line-height: 1.6;">
                    <strong>🔍 為什麼不只看證交所三大法人總表？</strong><br>
                    證交所每日公布的「外資買超」是將所有外資席位加總，無法分辨「波段真外資（如摩根士丹利、高盛）」與「隔日沖量化外資（如美林）」。<br>
                    透過本報告之分點穿透，能一眼辨認主力席位真實成本與鎖碼意志，避開隔日早盤倒貨陷阱！
                </div>

                <!-- 區塊 1: 外資席位重押榜 -->
                <div style="margin-bottom: 26px; border: 1px solid #cbd5e1; border-radius: 8px; overflow: hidden;">
                    <div style="background-color: #f8fafc; padding: 12px 16px; border-bottom: 2px solid #1e3a8a;">
                        <span style="font-size: 15px; font-weight: 800; color: #0f172a;">🌐 外資各大席位當日重押排行榜 (TOP 10)</span>
                        <span style="font-size: 11px; color: #64748b; margin-left: 8px;">(拆解摩根士丹利、摩根大通、高盛、美林、瑞銀等)</span>
                    </div>
                    <div style="overflow-x: auto; padding: 4px 12px 10px;">
                        <table style="width: 100%; border-collapse: collapse; text-align: left;">
                            <thead>
                                <tr style="border-bottom: 2px solid #cbd5e1; font-size: 11px; color: #475569; height: 32px;">
                                    <th style="padding: 6px 8px;">外資席位分點</th>
                                    <th style="padding: 6px 8px;">重押股票標的</th>
                                    <th style="padding: 6px 8px; text-align: right;">單日淨買超</th>
                                    <th style="padding: 6px 8px; text-align: right;">買進均價</th>
                                    <th style="padding: 6px 8px; text-align: right;">買進純度</th>
                                    <th style="padding: 6px 8px; text-align: center;">主力屬性</th>
                                </tr>
                            </thead>
                            <tbody>
                                {foreign_rows}
                            </tbody>
                        </table>
                    </div>
                </div>

                <!-- 區塊 2: 本土法人與總公司專戶重押榜 -->
                <div style="margin-bottom: 26px; border: 1px solid #cbd5e1; border-radius: 8px; overflow: hidden;">
                    <div style="background-color: #f8fafc; padding: 12px 16px; border-bottom: 2px solid #b45309;">
                        <span style="font-size: 15px; font-weight: 800; color: #0f172a;">🏢 本土法人業務部與總公司專戶重押榜 (TOP 10)</span>
                        <span style="font-size: 11px; color: #64748b; margin-left: 8px;">(投信經理人、代操資金與大戶專戶同步建倉標的)</span>
                    </div>
                    <div style="overflow-x: auto; padding: 4px 12px 10px;">
                        <table style="width: 100%; border-collapse: collapse; text-align: left;">
                            <thead>
                                <tr style="border-bottom: 2px solid #cbd5e1; font-size: 11px; color: #475569; height: 32px;">
                                    <th style="padding: 6px 8px;">法人券商分點</th>
                                    <th style="padding: 6px 8px;">重押股票標的</th>
                                    <th style="padding: 6px 8px; text-align: right;">單日淨買超</th>
                                    <th style="padding: 6px 8px; text-align: right;">買進均價</th>
                                    <th style="padding: 6px 8px; text-align: right;">買進純度</th>
                                    <th style="padding: 6px 8px; text-align: center;">吸籌特徵</th>
                                </tr>
                            </thead>
                            <tbody>
                                {inst_rows}
                            </tbody>
                        </table>
                    </div>
                </div>

                <!-- 區塊 3: 全市場分點多空之王 -->
                <div style="margin-bottom: 18px; border: 1px solid #cbd5e1; border-radius: 8px; overflow: hidden;">
                    <div style="background-color: #f8fafc; padding: 12px 16px; border-bottom: 2px solid #047857;">
                        <span style="font-size: 15px; font-weight: 800; color: #0f172a;">⚔️ 全市場分點多空之王排行 (今日最大主力買方 vs 賣方)</span>
                    </div>
                    <div style="overflow-x: auto; padding: 4px 12px 10px;">
                        <table style="width: 100%; border-collapse: collapse; text-align: left;">
                            <thead>
                                <tr style="border-bottom: 2px solid #cbd5e1; font-size: 11px; color: #475569; height: 32px;">
                                    <th style="padding: 6px 8px;">分點名稱與陣營</th>
                                    <th style="padding: 6px 8px; text-align: right;">全日淨買超/淨賣超</th>
                                    <th style="padding: 6px 8px; text-align: right;">總進出規模</th>
                                </tr>
                            </thead>
                            <tbody>
                                {kings_rows}
                            </tbody>
                        </table>
                    </div>
                </div>

                <!-- 說明區 -->
                <div style="padding: 12px 16px; background-color: #f8fafc; border-radius: 6px; font-size: 11px; color: #64748b; line-height: 1.5;">
                    💡 <strong>使用說明</strong>：本報表隨信附上包含完整明細之 Excel 檔案（含 3 個工作表）。若欲追蹤單一標的分點持續天期，請參閱每日主力四週期重押日報。
                </div>

            </div>

            <!-- Footer -->
            <div style="background-color: #f8fafc; padding: 14px 24px; text-align: center; font-size: 11px; color: #94a3b8; border-top: 1px solid #e2e8f0;">
                台股券商分點量化分析系統自動產製 ｜ 本日報僅供量化數據研究，不構成任何投資買賣建議。
            </div>

        </div>
    </body>
    </html>
    """

    subject = f"🏛️ 台股外資與本土法人分點買賣超解密日報 ({actual_date}) | 大摩高盛重押 ＋ 投信法人部鎖碼"
    attachments = [output_excel] if output_excel and os.path.exists(output_excel) else None

    success = send_email_report(
        subject=subject,
        html_content=full_html,
        recipients=recipients,
        attachment_paths=attachments
    )

    if success:
        tg_msg = (
            f"🏛️ *台股外資與本土法人分點買賣超日報 ({actual_date})*\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🌐 外資主要席位重押：`{len(df_foreign)} 檔`\n"
            f"🏢 本土法人部同步鎖碼：`{len(df_inst)} 檔`\n"
            f"👑 多頭第一名：`{df_market_kings.iloc[0]['券商分點']}` (+{df_market_kings.iloc[0]['net_amt_yi']:.1f}億)\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📧 專屬 HTML 郵件與 3-Sheet Excel 報表已成功寄出！"
        )
        send_telegram_notify(tg_msg)

    return success


def print_terminal_summary(df_foreign: pd.DataFrame, df_inst: pd.DataFrame, df_kings: pd.DataFrame):
    """於終端印出格式化繁體中文摘要"""
    if not df_foreign.empty:
        print("\n" + "=" * 90)
        print("🌐 【外資主要席位重押排行 TOP 10】")
        print("-" * 90)
        print(f"{'外資分點':<14} {'代號':<6} {'股票名稱':<8} {'淨買超(億)':>10} {'買均價':>8} {'純度%':>6} {'外資屬性':<14}")
        print("-" * 90)
        for _, r in df_foreign.head(10).iterrows():
            print(f"{str(r['券商分點'])[:10]:<14} {str(r['股票代號']):<6} {str(r['股票名稱'])[:6]:<8} {r['net_amt_yi']:>10.2f} {r['buy_avg_price']:>8.1f} {r['buy_purity_pct']:>6.1f} {str(r['外資屬性']):<14}")

    if not df_inst.empty:
        print("\n" + "=" * 90)
        print("🏢 【本土法人業務部與總公司專戶重押排行 TOP 10】")
        print("-" * 90)
        print(f"{'法人分點':<14} {'代號':<6} {'股票名稱':<8} {'淨買超(億)':>10} {'買均價':>8} {'純度%':>6} {'鎖碼特徵':<14}")
        print("-" * 90)
        for _, r in df_inst.head(10).iterrows():
            print(f"{str(r['券商分點'])[:10]:<14} {str(r['股票代號']):<6} {str(r['股票名稱'])[:6]:<8} {r['net_amt_yi']:>10.2f} {r['buy_avg_price']:>8.1f} {r['buy_purity_pct']:>6.1f} {str(r['法人標籤']):<14}")
    print("=" * 90 + "\n")


def main():
    parser = argparse.ArgumentParser(description="外資與本土法人分點買賣超解密引擎")
    parser.add_argument("--data-dir", default=r"d:\MyProject\stock_data_analysis\20260822分點資料", help="Parquet 資料目錄")
    parser.add_argument("--date", default=None, help="指定分析日期 (YYYY-MM-DD，預設取最新日期)")
    parser.add_argument("--min-foreign-amt", type=float, default=0.5, help="外資單股淨買超門檻 (億元，預設 0.5 億)")
    parser.add_argument("--min-inst-amt", type=float, default=0.2, help="本土法人單股淨買超門檻 (億元，預設 0.2 億)")
    parser.add_argument("--output-excel", default=None, help="匯出 Excel 報表路徑 (選填)")
    parser.add_argument("--send-email", action="store_true", help="產出後自動發送 Email 報告")

    args = parser.parse_args()

    df_foreign, df_inst, df_market_kings, actual_date = run_institutional_ranking_analysis(
        data_dir=args.data_dir,
        target_date=args.date,
        min_foreign_net_amt_yi=args.min_foreign_amt,
        min_inst_net_amt_yi=args.min_inst_amt
    )

    if not df_foreign.empty or not df_inst.empty:
        print_terminal_summary(df_foreign, df_inst, df_market_kings)

        output_excel = args.output_excel
        if not output_excel:
            output_excel = os.path.join(
                os.path.dirname(__file__),
                "output",
                f"外資與本土法人分點買賣超解密日報_{actual_date}.xlsx"
            )

        saved_excel = export_institutional_rankings_to_excel(
            df_foreign, df_inst, df_market_kings, output_excel
        )

        if args.send_email:
            send_institutional_rankings_email(
                df_foreign=df_foreign,
                df_inst=df_inst,
                df_market_kings=df_market_kings,
                actual_date=actual_date,
                output_excel=saved_excel
            )


if __name__ == "__main__":
    main()
